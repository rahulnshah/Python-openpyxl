from openpyxl import Workbook
# when running this script, do not have test.xlsx open, open it after running this script to see the changes
workbook = Workbook()
# sheet = workbook.active

worksheet1 = workbook.create_sheet("Mysheet")
worksheet1['A1'] = "h"
worksheet1['A2'] = "j"
worksheet1['A3'] = "c"

worksheet1['B1'] = "h"
worksheet1['B2'] = "j"
worksheet1['B3'] = "c"
worksheet1['B4'] = "i"

worksheet1['C1'] = "h"
worksheet1['C2'] = "j"
worksheet1['C3'] = "c"
worksheet1['C4'] = "t"
worksheet1['C5'] = "x"


def copy_pattern_into_rows(colLetter, ws, startNum, endNum, numOfTimes = 1):
    # llop through cells from colLetter + str(startNum) to colLetter + str(endNum - 1)
    nextEmptyCellNum  = endNum 
    for time in range(numOfTimes):
        for cell in range(startNum, endNum):
            ws[colLetter + str(nextEmptyCellNum)] = ws[colLetter + str(cell)].value
            nextEmptyCellNum = nextEmptyCellNum + 1
              
    
theColumns = ['A', 'B', 'C']
for letter in theColumns:
    sN = input("Enter starting cell row number: ")
    eN = input("Enter ending cell row number: ")
    copy_pattern_into_rows(letter, worksheet1, int(sN), int(eN))

copy_pattern_into_rows('C', worksheet1, 1, 6, 2)

    
    
# print(copy_pattern_into_rows("A", worksheet1, 1, 3, 2))
workbook.save('test.xlsx')

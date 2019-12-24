
#Import openyxl pythons xlsx reader
import openpyxl

#load xlsxformula into variable
wb=openpyxl.load_workbook('peter1.xlsx')

#load evaluated data into variable
wbeval=openpyxl.load_workbook('peter1.xlsx', data_only=True)

#list sheet names on workbook
print(wb.sheetnames)

# setup sheets in variables
sheet=wb['Sheet1']
sheeteval=wbeval['Sheet1']

# setup cell data into variables
a=sheet['c7']
b=sheet['e10']
b2=sheeteval['e10']

#print variables (note formula and evaluated)
print(a.value, b.value,b2.value)
